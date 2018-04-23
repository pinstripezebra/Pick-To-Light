# -*- coding: utf-8 -*-
"""
Created on Thu Apr 19 13:20:14 2018

@author: Lucas.See
"""
from openpyxl import load_workbook
import os
import xlrd

'''Generates a two dimensional array holding the values found in the Query tab of
the Scheduler'''
def parse_query():
    columns = 7
    units_day = 11
    rows = 0          
    wb = xlrd.open_workbook('C:\\Users\\seelc\\OneDrive\\Desktop\\Python\\Advanced Python\\Scheduler sample data.xlsx')
    sheet = wb.sheet_by_name('Query')
    sheet_2 = wb.sheet_by_name('Sheet2')
    data_storage = list()
    #for i in range(rows):
    #i = rows
    print('start')
    #Determining number of non empty rows in "Query"
    while(True):
        try:
            if sheet.cell(rows,1).value != '':
                rows = rows + 1
            print(rows)
        except:
            print("Ran out of rows")
            break
        
    print(rows)      
    for i in range(rows):    
        column = list()
        for j in range(columns):
            current_cell = sheet.cell(i, j)
            column.append(current_cell.value)
        data_storage.append(column)
        print(sheet.cell(i,1).value)
    print("Done_1")
    return data_storage

'''Takes data imported from sequencer and returns a dictionary populated with 
(order number): {option number} keys'''
def generate_units(data_storage):
    #Can find matches based off of order number
    option_storage = {}
    print(option_storage.keys)
    print(option_storage.keys())
    for i in range(1,rows):
    
        #if the order has already been added to order_keys
        if (data_storage[i][5]) in option_storage.keys():
            print("updating")
            #print(option_storage[data_storage[i][5]])
            option_storage[data_storage[i][5]].append(data_storage[i][1])
        
        #if the order number hasnt been added to storage matrix yet
        else:
            print("adding")
            option_storage.update({data_storage[i][5]: [data_storage[i][1]]})
    return option_storage

'''Takes dictionary containing all order number, line number combinations and
returns dictionary containing all order numbers as keys, merges line numbers'''            
def sort_unique_orders(option_storage):
        
    #Merging orders with common line numbers
    unique_orders = {}
    my_keys = list(option_storage.keys())
    unique_keys = []

    #Creating a list of all keys regardless of line number
    for keys in my_keys:
        split_key = keys.split('^')
        unique_key = split_key[0]
        if unique_key not in unique_keys:
            unique_keys.append(unique_key)

    #Creating a new dictionary with the unique keys
    for keys in unique_keys:
        part_numbers = list()
        for all_keys in my_keys:
            if all_keys.find(keys) != -1:
                part_numbers.append(option_storage[all_keys])
        unique_orders[keys] = part_numbers
    print("done")
    return unique_orders

'''Takes the number type, ex. "GT" and the starting dictionary and returns a
filtered dictionary with only the chosen number types as values'''
def filter_pn(number_type, orders):
    filtered_orders = {}
    for keys in orders:
        print(keys)
        filtered_numbers = list()
        for number in orders[keys]:
            for item in number:
                if item.find(number_type) != -1:
                    #print("Entered")
                    filtered_numbers.append(item)
        filtered_orders[keys] = filtered_numbers
    return filtered_orders
                
       
if __name__ == "__main__":
    intake = parse_query()
    all_units = generate_units(intake)
    unique_units = sort_unique_orders(all_units)
    usable_dict = filter_pn('GT', unique_units)
    

